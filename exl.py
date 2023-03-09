import pandas as pd
from openpyxl import load_workbook
import os

# pobieranie arkusza spedytora i poprawnie podstawy do obliczenia prowizji jeżeli coś się nie zgadza
def add(z):

    teams = z.drop_duplicates("ID Teamu")["ID Teamu"].size
    i=0
    skladowa = 0.0
    sumapd = 0.0

    while(i!=3):
        print("Potracenia/dodatki")
        print("1 - Przejdz przez liste wszystkich spedytorow ")
        print("2 - Wybierz konkretnego spedytora ")
        print("3 - zakoncz")
        print("\n")

        i = int(input("wybór :"))

        if i == 1 :
            for i in range(1, teams + 1):

                for j in range(0, z[z["ID Teamu"] == i]["ID Teamu"].size):

                    sumapd = 0
                    k = 1
                    while (k != 0):
                        print(list(z[z["ID Teamu"] == i]["Spedytor"])[j])
                        print("Dodaj potracenie/dodatek : ")
                        skladowa = float(input())
                        sumapd += skladowa
                        print("\n")
                        print('suma :')
                        z.at[j, "potracenia/dodatki"] = sumapd
                        print(sumapd)
                        print("\n")
                        print(
                            "Jezeli chcesz dodac kolejny koszt/potracenie dla tego spedytora nacisnij dowolna cyfre JEZELI NIE nacisnij 0")
                        k = int(input())
                        print(z["potracenia/dodatki"])
                        print("\n")

        if i == 2:



            id = input("Wprowadz ID spedytora : ")
            k = 1
            while (k != 0):
                print(z[z.index == id])
                print("Dodaj potracenie/dodatek : ")
                skladowa = float(input())
                sumapd += skladowa
                print("\n")
                print('suma :')
                print(sumapd)
                print("\n")
                print(
                    "Jezeli chcesz dodac kolejny koszt/potracenie dla tego spedytora nacisnij dowolna cyfre JEZELI NIE nacisnij 0")
                k = int(input())
                print("\n")


    return z




def choice_path():


    #os.chdir()

    path = []

    folder_path = 'arkusze'



    for filename in os.listdir(folder_path):

        file_path = os.path.join(folder_path,filename)
        if os.path.isfile(file_path):

            print('Plik:', filename)
            path.append('arkusze/'+filename)
            wb = load_workbook(filename='arkusze/'+filename)
            # UWAGA !!!!
            # ARKUSZE MUSZĄ BYĆ USTAWIONĘ W NASTĘPUJĄCEJ KOLEJNOŚCI
            # RAPORT SPEDYTORZY -> RAPORT SPEDYTOR -> STAWKI PROWIZJI
            path.append(wb.sheetnames[0])
            path.append(wb.sheetnames[1])
            path.append(wb.sheetnames[2])

        elif os.path.isdir(file_path):
            print('Folder:', filename)


    return path


def check(x, y, name, procedure):

    #musisz wykonać tekstową procedure
    

    podstawa = 0

    for i in range(0, x["Klient"].size - 1):
        if name in x["Klient"][i]:
         podstawa += procedure(int(x["Prowizja"][i]))
        else:
         podstawa += int(x["Prowizja"][i])


    return podstawa




def count_commission(z,stawki):


    teams = z.drop_duplicates("ID Teamu")["ID Teamu"].size

    S = lambda wynik, stawka,sumapd: wynik * stawka + sumapd
    T2 = lambda wynik, stawka,sumapd: (wynik * stawka) * 0.82 + sumapd
    T3 = lambda wynik, stawka,sumapd, koszty: (wynik - koszty) * stawka + sumapd
    p=1
    for i in range(1, teams+1):


        for j in range(0, z[z["ID Teamu"] == i]["ID Teamu"].size):

            prow = list(z[z["ID Teamu"] == i]["Prowizja"])[j]
            prog = list(stawki[stawki.index == list(z[z["ID Teamu"] == i]["ID Spedytora"])[j]]["Próg PLN"])
            #print(stawki[stawki.index == list(z[z["ID Teamu"] == i]["ID Spedytora"])[j]]["Próg PLN"])
            wyplata = 0
            stopa = 0
            koszty = 0
            sumapd = list(z[z["ID Teamu"] == i]["dodatki"])[j] - list(z[z["ID Teamu"] == i]["potracenia"])[j]

            if ("S" in list(z[z["ID Teamu"] == i]["ID kalkulacji"])[j]):
                for k in range(0, len(prog)):

                    if (prog[k] - prow) < 0:

                        stopa = list(stawki[stawki.index == list(z[z["ID Teamu"] == i]["ID Spedytora"])[j]]["Stawka %"])[k]




                wyplata = S(prow, stopa, sumapd)


            if ("T" in list(z[z["ID Teamu"] == i]["ID kalkulacji"])[j]):

                prow = z[z["ID Teamu"] == i]["Prowizja"].sum()

                #print(list(z[z["ID Teamu"] == i]["ID Spedytora"])[j])
                for k in range(0, len(prog)):

                    if (prog[k] - prow) < 0:

                        stopa = list(stawki[stawki.index == list(z[z["ID Teamu"] == i]["ID Spedytora"])[j]]["Stawka %"])[k]
                        #print(stopa)
                        #print("*****")

                if ("T1" in list(z[z["ID Teamu"] == i]["ID kalkulacji"])[j]):

                    wyplata = S(prow, stopa, sumapd)
                if ("T2" in list(z[z["ID Teamu"] == i]["ID kalkulacji"])[j]):
                    wyplata = T2(prow, stopa, sumapd)
                if ("T3" in list(z[z["ID Teamu"] == i]["ID kalkulacji"])[j]):

                    wyplata = T3(prow, stopa,sumapd,list(z[z["ID Teamu"] == i]["koszty"])[j])


            d = {'Spedytor': [list(z[z["ID Teamu"] == i]["Spedytor"])[j]] ,'ID teamu': [list(z[z["ID Teamu"] == i]["ID Teamu"])[j]],'ID': [list(z[z["ID Teamu"] == i]["ID Spedytora"])[j]],'id kalkulacji': [list(z[z["ID Teamu"] == i]["ID kalkulacji"])[j]],'procent': stopa ,'wynik': prow, 'wyplata': [wyplata]}
            df = pd.DataFrame(data=d).set_index('ID')



            with pd.ExcelWriter("arkusze/nazwa.xlsx", mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer :
                df.to_excel(writer,sheet_name="wyplaty",startrow=(p))
                p+=2
            print(">>>")

            #print(z[z["ID Teamu"] == i])

